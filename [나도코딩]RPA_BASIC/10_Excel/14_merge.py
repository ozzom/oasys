from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기 
ws.merge_cells("B2:D2")  #B2부터 D2 병합 
ws["B2"].value = " 병합된 셀입니다. "

wb.save("sample_merge.xlsx")

