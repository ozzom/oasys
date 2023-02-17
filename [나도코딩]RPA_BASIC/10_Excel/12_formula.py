import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = datetime.date.today()  # 오늘 날짜 정보
ws["A2"] = "=sum(1, 2, 3)"        # 1 + 2 + 3 = 6
ws["A3"] = "=AVERAGE(1, 2, 3)"    

ws["B1"] = 30
ws["B2"] = 20
ws["B3"] = "=sum(b1:b2)"    

wb.save("sample_formula.xlsx") 