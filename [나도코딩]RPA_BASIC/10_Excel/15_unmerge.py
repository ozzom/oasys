from openpyxl import load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

#셀 합치기 
ws.unmerge_cells("B2:D2")  #B2부터 D2다시 분할 
wb.save("sample_unmerge.xlsx")
