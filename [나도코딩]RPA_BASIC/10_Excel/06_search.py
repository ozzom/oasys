from openpyxl import load_workbook

wb =load_workbook("5_cell_range.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학 순으로 가져올 것인데.. 
    if row[1].value > 80:
        print(f"{row[0].value}번 학생은 영어성적이 훌륭합니다.({row[1].value}점)")
        

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("5_cell_range_modified.xlsx")            
wb.save("5_cell_range_modified.xlsx.dprtpf")         