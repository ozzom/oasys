from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

#번호 (국어) 영어 수학 
# ws.move_range("b1:c11", rows=0, cols=1)  #  move_range("이동할대상", rows=0, cols=1) rows, cols 는 B1 기준으로 이동  
# ws["b1"].value = "국어"

ws.move_range("c1:c11", rows=5, cols=-1)  # C1 기준으로 이동
# wb.save("sample_korean.xlsx")

wb.save("sample_korean.dprtpf")
# wb.save("sample_korean.xlsx.dprtpf")

 