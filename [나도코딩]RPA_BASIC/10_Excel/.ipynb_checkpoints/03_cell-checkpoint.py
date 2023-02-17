from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws.title = "WenTaiXiongSheet"

# 셀에 데이터 쓰기 
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["b1"] = 4 
ws["b2"] = 5
ws["b6"] = 6

#셀에 있는 데이터 읽기 
print(ws["a1"])                     # A1 셀의 정보를 출력 result : <Cell 'WenTaiXiongSheet'.A1>
print(ws["a1"].value)               # A1 셀의 값을 출력   result : 1
print(ws["B1"].value)

# row = 1, 2, 3, 
# column = A(1), B(2), C(3), .... 
print(ws.cell(row=1, column=1))           # 매개변수를 선언하는 방식은 row, column을 바꿔도 무관 
print(ws.cell(row=1, column=1).value)     # ws["A1"].value
print(ws.cell(row=1, column=2).value)     # ws["B1"].vlaue 

print(ws.cell(1, 1)) 
print(ws.cell(1, 1).value) 
print(ws.cell(1, 2).value) 


ws.cell(column=3, row=1, value=10)  # ws["C3"] = 10 or ws["C3"].value = 10 과 동일  
print(ws.cell(1,3).value)

from random import *
#반복문을 통해서 랜덤 숫자 채우기 

index = 1 

for x in range(1, 11):     # 10개 row 채우기 
    for y in range(1, 11): # 10개 Col 채우기 
        # ws.cell(row=x, column=y, value=randint(1, 100))  #1~100사이의 숫자 채우기 
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("3_cell.xlsx.dprtpf")

  