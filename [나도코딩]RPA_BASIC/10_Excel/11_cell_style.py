from openpyxl import load_workbook
from openpyxl.styles import Font #폰트
from openpyxl.styles import Border  #외곽선 
from openpyxl.styles import Side
from openpyxl.styles.borders import Side
from openpyxl.styles import PatternFill #패턴으로 채우기 
from openpyxl.styles import Alignment  #정렬


wb = load_workbook("sample.xlsx")
ws = wb.active

# #번호, 영어, 수학 각 셀에 스타일 적용
a1 = ws["A1"]  # 번호 
b1 = ws["B1"]  # 영어
c1 = ws["C1"]  # 수학 

# # A열의 너비를 5로 설정 
ws.column_dimensions["A"].width = 5

# # 1행의 높이를  50로 설정 
ws.row_dimensions[1].height = 50

# 스타일 적용
a1.font = Font(color="FF0000", italic=True, bold=True)    # 글자색 빨강, 이탤랙, 볼드 적용
b1.font = Font(color="CC33ff", name="Arial", strike=True) # name 글꼴, strike 취소선
c1.font = Font(color="0000ff", size=20, underline="single") # 글자 크기 20, 밑줄 적용(한줄-single, 두줄:double)

# 테두리 적용 
# 왼쪽, 오른쪽, 위쪽, 아래에 대해서 테두리 적용하는 객체 생성  
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))  
a1.border = thin_border   # 셀에 테두리 적용 
b1.border = thin_border   # 셀에 테두리 적용 
c1.border = thin_border   # 셀에 테두리 적용 

# 셀의 배경색 변경 
# 90 점 넘는 셀에 대해서 초록색 적용 
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center") #center, left, right, top, bottom
        if cell.column == 1: #A 번호열은 제외 
            continue
        
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00" , fill_type="solid")
            cell.font = Font(color="FF0000")
            
# 틀 고정 
ws.freeze_panes = "B2" #B2 기준으로 틀 고정 
               
wb.save("sample_chart.xlsx")

