from openpyxl import Workbook
from openpyxl import load_workbook  # 파일 불러오기 
wb = load_workbook("3_cell.xlsx")
ws = wb.active      # 활성화된 sheet 

# 셀 데이터 불러오기 -> 셀 갯수를 알고 있을 때는 명시적으로 접근 
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()
    
# 셀 갯수를 모를 때는 최대 칼럼 값을 이용해서 접근 
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
    