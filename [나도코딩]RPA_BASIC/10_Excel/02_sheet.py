from openpyxl import Workbook

wb = Workbook()                          # 엑셀파일로 생각하면 된다. 
ws = wb.create_sheet()                   # 새로운 sheet 기본이름으로 생성 
ws.title = "MySheet"                     # sheet 이름이 변경 됨 
ws.sheet_properties.tabColor = "ff66ff"  # rgb 형태로 값을 넣어주면 탭 색상 변경 

#sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")    # 주어진 이름으로 Sheet 생성, 기본은 마지막에 sheet가 생성 됨 
ws2 = wb.create_sheet("NewSheet", 2)  # 2번째 인덱스에 Sheet 생성 

# WorkBook에서 Sheet 명으로 접근하기 
new_ws = wb["NewSheet"]  # dict 형태로 Sheet에 접근 

print(wb.sheetnames)   # 모든 Sheet 이름 확인 

#new_ws 시트 내용을 Target_ws시트에 복사 
new_ws["A1"]= "test"        # new_WS의 A:1에 test를 입력 
target_ws = wb.copy_worksheet(new_ws)
target_ws.title = "copied_sheet"

print(wb.sheetnames)   # 모든 Sheet 이름 확인 