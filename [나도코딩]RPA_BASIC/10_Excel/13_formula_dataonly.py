from openpyxl import load_workbook

# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active

#수식 그대로 가져오고 있음. 

# for row in ws.values:
#     for cell in row:
#         print(cell) 
        
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# 수식이 아니라 데이터만 가져오는 방법        
# evaluate 되지 않은 상태의 데이터는 None 이라고 표시 
# 파일을 한번 열어보고 저장한 후에는 정상적으로 읽을 수 있음. 
for row in ws.values:
    for cell in row: 
        print(cell) 
