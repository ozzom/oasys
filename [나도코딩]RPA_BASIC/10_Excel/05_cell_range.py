from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import rows_from_range 

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기 
ws.append(["번호","영어","수학"])    # 제일 첫줄에 Header를 넣는 것. "번호" A칼럼,"영어" B 칼럼, "수학" C 칼럼 
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(1, 100)])

# col_B = ws["B"]  # 영어 칼럼만 가져오려고 할 때, 
# print(col_B)
# for cell in col_B:
#     print(cell.value)
    
# col_range = ws["B:C"]  # 영어, 수학 열 함께 가져오기 
# for cols in col_range: # B 열을 먼저 Loop 돌고, C칼럼을 Loop를 돌아서 tuple 로 Return 
#     for cell in cols:
#         print(cell.value)

# row_title = ws[1]     # 첫번째 row 만 가져오기 
# for cell in row_title:
#     print(cell.value)
    
# row_range = ws[2:6]   # 첫번째 줄인 Title을 제외하고, 2번째 줄에서 6번째 줄까지 가지고 오기 
# for row in row_range: # 한 줄을 가져온다. 
#     for cell in row:  # 위에서 가져올 줄 에서 Cell 하나씩 가져온다. 
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string
    
# row_range = ws[2:ws.max_row]  # 범위를 모르니 전부 가져온다. 
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ")   # 셀의 좌표 정보를 가져옴, A10, AZ250 
#         xy = coordinate_from_string(cell.coordinate)  #셀의 좌표 정보를 ('A', 2) 형태로 가져옴 
#         # print(xy, end=" ")
#         print(xy[0], end="")     # 셀의 칼럼 좌표값 정보만 가져옴  A B C 
#         print(xy[1], end=" ")    # 셀의 row 좌표값                1, 2, 3, 4, ... ,10
#     print()
    
# wb.save("5_cell_range.xlsx.dprtpf")    

# 전체 행을 가져오고 싶을 때, 
# print(tuple(ws.rows)) # 한줄 씩, 튜블로 묶어서 가져옴, 헤더, 1번 줄, 2번 줄,.... , 10줄로 11개의 튜플을 리턴 
# for row in tuple(ws.rows):
#     print(row[0].value)   # Cell 값을 가져오는 칼럼은 Index 이고, 줄은 자동으로 증가한다. ==> 특정 열의 모든 값을 가져오는 것 

# for row in ws.iter_rows():
#     print(row[0].value )

# 전체 행을 가져온다. min_row, max_row, min_col, max_col 을 이용해서 슬라이싱 하는 것과 같음 
# 데이터를 좌우좌우로 가져오낟. 
for row in ws.iter_rows(min_row=2, max_row=11, min_col = 2, max_col= 3): 
    # print(row[0])
    print(row[0].value, row[1].value) #  영어, 수학 점수 
    
    
# print(tuple(ws.columns)) # 한 칼럼씩, 튜블로 묶어서 가져옴, 번호로 하나, 영어로 하나, 수학으로 하나 3개의 튜플을 가져옴 
# for col in tuple(ws.columns):
#     print(col[2].value)

# for col in ws.iter_cols():
#     print(col[0].value )

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):  # 상하 상하로 데이터를 가져옴 
    print(col[0].value, col[1].value, col[2].value, col[3].value, col[4].value)
    
    


